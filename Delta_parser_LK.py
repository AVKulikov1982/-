import os
import time
import requests
import logging
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from prepare import prepare_name
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger('__name__')
dotenv_path = os.path.join(".env")
load_dotenv(dotenv_path=dotenv_path)
load_dotenv()

path_analitic = os.getenv("path_analitic")
HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:96.0) Gecko/20100101 Firefox/96.0'}
session = requests.Session()

HOST = 'https://deltasert.ru'  # хост
link = HOST + '/login'

data = {
	"login": os.getenv("login_delta"),  # логин пользователя
	"password": os.getenv("pass_delta"),  # пароль пользователя
	"form_action": "login",
	"enter": "Войти"
}

response = session.post(link, data=data, headers=HEADERS)
time.sleep(1)


def get_response_bills(url: str) -> requests.Response:
	"""
	Функция на вход получает url, возвращает Response объект.
	:param url: url
	:return:
	"""
	response_bills = session.get(url, headers=HEADERS)
	print(response_bills.status_code)
	return response_bills


def get_pages_count(soup) -> int:
	"""
	Функция — обработчик пагинации; на вход получает объект soup, возвращает количество страниц (пагинация).
	:param soup: object BeautifulSoup
	:return:
	"""
	pagination_to = soup.find('div', class_='PageSelector')
	pages_count = 1
	if pagination_to:
		pagination = pagination_to.find_all('a')
		last_href = pagination[-1].get('href')
		pages_count = int(last_href.split('&')[0].split('=')[-1])
	return pages_count


def parse():
	"""
	Функция парсер.
	:return:
	"""
	URL = HOST + '/client/bills'
	response_bills = get_response_bills(URL)  # получаем объект Response
	soup = BeautifulSoup(response_bills.text, 'html.parser')  # получаем объект BeautifulSoup
	# print(soup)
	pages_count = get_pages_count(soup)  # получаем количество страниц (пагинация)
	for page in range(1, pages_count + 1):
		logger.info(f'Парсинг страницы {page} {pages_count} {URL}...')
		html = session.get(URL, headers=HEADERS, params={'page': page})
		soup = BeautifulSoup(html.text, 'html.parser')
		bills = soup.find_all('a', class_='list_item')
		for bill in bills:
			bill_date = bill.find_all('div', class_='managers_lists_container_block')[1].get_text().replace(' ', '_')
			response_to_bill = session.get(HOST + bill.get('href'), headers=HEADERS)
			soup_bill = BeautifulSoup(response_to_bill.text, 'html.parser')
			list_links = soup_bill.find('div', id='page_buttons_container').find_all('a', class_='success')
			list_data = list(
				map(lambda x: x.get_text().strip(), soup_bill.find_all('div', class_='customer_overall_block_right')))
			time.sleep(1)
			for doc_link in list_links:
				path = doc_link.get('href').split('id=')[-1] + '-' + bill_date
				try:
					if list_data[4] != '0 ₽':
						if 'Счет на предоплату' in list_data:
							path += '_предоплата'
						if 'За счет предоплаты' in list_data:
							path += '_оплачено_предоплатой'
						if 'С бонусного счета' in list_data:
							path += '_оплачено_бонусами'
						if path not in os.listdir():
							os.mkdir(path)
					else:
						raise
					response_to_doc_link = session.get(HOST + doc_link.get('href'), headers=HEADERS)
					try:
						if 'счета' in doc_link.get_text():
							path += '/bill.pdf'
						elif 'детализации' in doc_link.get_text():
							path += '/detail.xlsx'
						elif 'акта' in doc_link.get_text():
							path += '/act.pdf'
						with open(path, 'wb') as f_o:
							f_o.write(response_to_doc_link.content)
						time.sleep(1)
					except:
						logger.error(f'Error parser')
				except:
					logger.error(f'Неоплаченный счет')


def get_info_leads():
	if not os.path.exists(path_analitic):
		create_excel()
	URL = HOST + '/client/requests'
	response_bills = get_response_bills(URL)  # получаем объект Response
	soup = BeautifulSoup(response_bills.text, 'html.parser')  # получаем объект BeautifulSoup
	# print(soup)
	pages_count = get_pages_count(soup)  # получаем количество страниц (пагинация)
	tmp_data = []
	for page in range(1, pages_count + 1):
		logger.info(f'Парсинг страницы {page} {pages_count} {URL}...')
		html = session.get(URL, headers=HEADERS, params={'page': page})
		leads = BeautifulSoup(html.text, 'html.parser').find_all('a', class_='managers_lists_container_block')
		for lead in leads:
			values_data = lead.find_all('div', class_='managers_lists_container_block')
			lead_data = []
			for i, v in enumerate(values_data):
				if i in [0, 1, 2, 6, 8, 9]:
					lead_data.append(v.text.strip())
				if i == 11:
					v = prepare_name(v.text)
					lead_data.append(v)
			tmp_data.append(lead_data)
	put_in_analitic(tmp_data)


def create_excel(excel_data=None):
	wb = Workbook()
	ws = wb['Сводная_Аналитик']
	for i, value in enumerate(
			['№ Заявки', 'Дата', 'Год', 'Документ', '№ Документа', 'Стоимость', 'Счет', 'Заявитель']):
		cell = ws.cell(column=i + 1, row=1, value=value)
		cell.font = Font(size=12, bold=True)
		cell.alignment = Alignment(horizontal="left", vertical="center")
		ws.column_dimensions[cell.column_letter].width = 15
	wb.remove(wb['Sheet'])
	wb.save(os.getcwd()+'/analitic_all.xlsx')


def put_in_analitic(data):
	wb = load_workbook(path_analitic, data_only=True)
	ws = wb.get_sheet_by_name('Сводная_Аналитик')
	for row, lead in enumerate(data, 2):
		for column, value in enumerate(lead, 1):
			cell = ws.cell(row=row, column=column, value=value)
	wb.save(path_analitic)


if __name__ == '__main__':
	logging.basicConfig(level=logging.INFO, filename='bot.log', filemode='a',
						format='%(asctime)s - %(levelname)s - %(message)s',
						datefmt='%d-%b-%y %H:%M:%S')
	logger.info(f'Start parser')
	# parse()
	# get_info_leads()
